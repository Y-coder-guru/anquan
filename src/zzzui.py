# -*- coding: utf-8 -*-
"""
智巡卫士 - Flask Web 版实验室智能安全监控系统

功能：环境监测、危险分析、人脸识别、报警日志、用户管理、系统设置、Web 视频流。
"""

import hashlib
import csv
import io
import json
import os
import pickle
import random
import re
import sys
import threading
import time
import traceback
import urllib.parse
from datetime import datetime, timedelta
import base64
import atexit

import requests
from flask import Flask, Response, jsonify, redirect, render_template, request, send_file, session, url_for

try:
    import cv2
    import numpy as np
except ModuleNotFoundError:
    cv2 = None
    np = None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FACE_DB_FILE = os.path.join(BASE_DIR, "lab_members.pkl")
USER_DB_FILE = os.path.join(BASE_DIR, "users.pkl")
APP_STATE_FILE = os.path.join(BASE_DIR, "app_state.json")
SNAPSHOT_DIR = os.path.join(BASE_DIR, "static", "snapshots")

TOLERANCE = 60
FACE_FEATURE_SIZE = (64, 64)
FACE_FEATURE_LENGTH = 32 * 32 + 256 + 16
FACE_MATCH_THRESHOLD = 0.32
FACE_AMBIGUITY_MARGIN = 0.06
ALARM_INTERVAL = 30
LOG_LIMIT = 300
CAMERA_INDEX = 0
PUSH_ENABLED = True
STATE_SAVE_INTERVAL = 8
SENSOR_HISTORY_INTERVAL = 30
SENSOR_HISTORY_LIMIT = 720
ASSET_VERSION = "20260521-auth-glass-tabs"

web_app = Flask(__name__, template_folder="templates", static_folder="static")
web_app.secret_key = "zhixunweishi_secret_key_2024"


user_database = {
    "admin": {
        "password": hashlib.md5("admin123".encode()).hexdigest(),
        "role": "实验室管理员",
        "email": "admin@lab.com",
        "display_name": "系统管理员",
        "department": "安全管理",
        "phone": "-",
        "register_time": "2026-05-19"
    }
}

DEFAULT_THRESHOLDS = {
    "room_temp_warn": 35,
    "smoke_alarm": 30,
    "voc_warn": 200,
    "gas_ch4_warn": 5000,
    "gas_ch4_explosion": 25000,
    "gas_co_alarm": 35,
    "gas_h2s_alarm": 20,
    "reactor_temp_alarm": 90,
    "reactor_pressure_alarm": 1.2
}

SENSOR_DEFINITIONS = [
    ("room_temp", "室温", "℃", 60, "room_temp_warn"),
    ("humidity", "湿度", "%", 100, None),
    ("smoke", "烟雾", "%", 100, "smoke_alarm"),
    ("voc", "VOC", "ppb", 2000, "voc_warn"),
    ("gas_ch4", "瓦斯 CH₄", "ppm", 50000, "gas_ch4_warn"),
    ("gas_h2", "氢气 H₂", "ppm", 40000, None),
    ("gas_co", "一氧化碳 CO", "ppm", 400, "gas_co_alarm"),
    ("gas_h2s", "硫化氢 H₂S", "ppm", 100, "gas_h2s_alarm"),
    ("reactor_temp", "反应釜温度", "℃", 120, "reactor_temp_alarm"),
    ("reactor_pressure", "反应釜压力", "MPa", 2, "reactor_pressure_alarm")
]

DEVICE_CATALOG = [
    {"id": "sensor_hub", "name": "环境传感器组", "category": "采集设备", "location": "实验室"},
    {"id": "ventilation", "name": "通风排风系统", "category": "联动设备", "location": "实验室顶部"},
    {"id": "suppression", "name": "灭火抑制系统", "category": "联动设备", "location": "实验室"},
    {"id": "access_control", "name": "门禁联动系统", "category": "安全设备", "location": "安全出口"},
    {"id": "camera", "name": "实时监控摄像头", "category": "视频设备", "location": "监控位"},
    {"id": "bark_push", "name": "Bark 手机推送", "category": "通知设备", "location": "iPhone"}
]

BARK_KEYS = {
    1: "pvzbjaksE24KDQffBbcNh",
    2: "",
    3: ""
}

PUSH_CONFIG = {
    1: {"title": "📢 一级预警", "targets": ["实验人员", "实验室管理员"], "key": BARK_KEYS[1]},
    2: {"title": "⚠️ 二级报警", "targets": ["实验室主任", "实验人员", "实验室管理员"], "key": BARK_KEYS[2]},
    3: {"title": "🚨 三级紧急", "targets": ["校领导", "保卫处", "实验室主任", "实验人员", "实验室管理员"],
        "key": BARK_KEYS[3]}
}

ACCIDENT_RULES = {
    "indoor_fire": {"name": "室内火灾", "level": 3, "conditions": {"smoke": {"min": 30}},
                    "fire_suppression": "🧯 吊顶式超细干粉灭火装置",
                    "actions": ["🔌 切断总电源", "🧯 启动自动灭火装置", "📢 紧急疏散广播", "🚪 自动打开安全门"],
                    "sequence": ["确认烟雾探测器报警并锁定火情区域", "切断实验台与非消防总电源",
                                 "启动吊顶式超细干粉灭火装置", "开启疏散广播并打开安全门",
                                 "通知实验室管理员与保卫处复核现场"]},
    "reactor_fire": {"name": "反应釜火灾", "level": 3, "conditions": {"reactor_temp": {"min": 90}},
                     "fire_suppression": "🧯 惰性气体抑制 + 夹套冷却",
                     "actions": ["❄️ 启动夹套冷却", "💨 注入惰性气体", "🔒 关闭进料阀", "🔧 打开泄压阀"],
                     "sequence": ["停止加热并关闭进料阀", "启动反应釜夹套冷却系统",
                                  "注入惰性气体进行火源抑制", "压力继续升高时打开泄压阀",
                                  "保留趋势数据并通知实验室主任"]},
    "gas_ch4_explosion": {"name": "瓦斯爆炸风险", "level": 3, "conditions": {"gas_ch4": {"min": 25000}},
                          "fire_suppression": "🧯 惰性气体抑制",
                          "actions": ["💨 排风扇最大", "🔒 关闭气源", "🔌 切断非防爆电源", "📢 疏散广播"],
                          "sequence": ["立即关闭燃气与可燃气体总阀", "切断非防爆电源，禁止火花源",
                                       "启动防爆排风至最大档位", "广播疏散并封控实验室入口",
                                       "等待浓度降至安全区间后人工复核"]},
    "h2s_leak": {"name": "硫化氢泄漏", "level": 3, "conditions": {"gas_h2s": {"min": 20}},
                 "fire_suppression": None, "actions": ["💨 最大排风", "🧪 喷淋中和", "📢 紧急疏散"],
                 "sequence": ["启动最大排风并维持负压", "触发有毒气体声光报警",
                              "开启喷淋或中和处置设备", "疏散人员并佩戴防护装备复核",
                              "记录泄漏源和处置闭环"]},
    "reactor_overpressure": {"name": "反应釜超压", "level": 2, "conditions": {"reactor_pressure": {"min": 1.2}},
                             "fire_suppression": None, "actions": ["🔧 打开泄压阀", "❄️ 启动冷却", "🔊 声光报警"],
                             "sequence": ["触发声光报警并锁定反应釜编号", "停止加热并启动冷却",
                                          "按安全阈值打开泄压阀", "通知值班人员现场确认",
                                          "压力恢复后保留报警记录"]},
    "co_leak": {"name": "一氧化碳泄漏", "level": 2, "conditions": {"gas_co": {"min": 35}},
                "fire_suppression": None, "actions": ["💨 强制排风", "🔒 关闭气源", "🔊 声光报警"],
                "sequence": ["启动一氧化碳声光报警", "强制排风并打开补风通道",
                             "关闭相关气源阀门", "提醒人员撤离并避免二次进入",
                             "浓度恢复后进行传感器复位"]},
    "gas_ch4_leak": {"name": "瓦斯泄漏", "level": 1, "conditions": {"gas_ch4": {"min": 5000}},
                     "fire_suppression": None, "actions": ["💨 启动防爆排风", "🔒 关闭气源", "🔊 预警提示"],
                     "sequence": ["发出一级预警并持续监测浓度", "启动防爆排风",
                                  "关闭疑似泄漏气源", "提示人员检查软管和阀门",
                                  "浓度继续升高则升级为爆炸风险"]},
    "high_temp_warning": {"name": "温度异常", "level": 1, "conditions": {"room_temp": {"min": 35}},
                          "fire_suppression": None, "actions": ["💨 启动排风扇", "🔊 提示音"],
                          "sequence": ["提示室温异常并记录趋势", "启动排风扇或空调降温",
                                       "检查热源设备是否异常", "通知实验人员确认现场",
                                       "温度恢复后自动解除预警"]}
}

NORMAL_SEQUENCE = ["持续采集环境与气体传感器", "保持联动设备待命", "人脸门禁持续巡检", "异常触发后自动生成处置序列"]

monitor_data = {
    "room_temp": 25.0, "humidity": 45, "smoke": 0, "voc": 50,
    "reactor_temp": 45, "reactor_pressure": 0.5,
    "gas_ch4": 0, "gas_h2": 0, "gas_co": 0, "gas_h2s": 0,
    "risk_level": 0, "risk_text": "✅ 系统正常", "risk_reason": "所有参数正常",
    "accident_type": "无", "alert_detail": "", "actions": [],
    "alert_history": [], "action_history": [], "face_count": 0, "current_faces": 0,
    "camera_running": False, "camera_error": "", "push_targets": [],
    "emergency_sequence": NORMAL_SEQUENCE[:], "fire_suppression": None,
    "stranger_alarm_count": 0, "known_face_count": 0, "unknown_face_count": 0,
    "last_face_event": "暂无人脸事件", "last_stranger_time": "--", "detected_faces": []
}

face_members = {}
last_alarm_time = {}
system_logs = []
sensor_history = []
state_dirty = False
last_state_save = 0
service_started = False
data_lock = threading.RLock()


def safe_print(text=""):
    try:
        if sys.stdout is None:
            return
        encoding = sys.stdout.encoding or "utf-8"
        safe_text = str(text).encode(encoding, errors="replace").decode(encoding)
        sys.stdout.write(safe_text + "\n")
        sys.stdout.flush()
    except Exception:
        pass


def add_log(message):
    entry = {"time": current_time_text(), "message": message}
    with data_lock:
        system_logs.insert(0, entry)
        del system_logs[LOG_LIMIT:]
        mark_state_dirty()
    safe_print(f"[{entry['time']}] {message}")
    if service_started:
        save_app_state(force=True)


def load_user_database():
    global user_database
    if os.path.exists(USER_DB_FILE):
        try:
            with open(USER_DB_FILE, "rb") as f:
                loaded_users = pickle.load(f)
            if isinstance(loaded_users, dict):
                user_database.update(loaded_users)
        except Exception as exc:
            add_log(f"⚠️ 用户数据库读取失败: {exc}")
    user_database.setdefault("admin", {
        "password": hashlib.md5("admin123".encode()).hexdigest(),
        "role": "实验室管理员",
        "email": "admin@lab.com",
        "display_name": "系统管理员",
        "department": "安全管理",
        "phone": "-",
        "register_time": "2026-05-19"
    })
    for username, info in user_database.items():
        info.setdefault("display_name", username)
        info.setdefault("department", "-")
        info.setdefault("phone", "-")
        info.setdefault("email", "-")
        info.setdefault("role", "实验人员")
        info.setdefault("register_time", "-")


def save_user_database():
    with open(USER_DB_FILE, "wb") as f:
        pickle.dump(user_database, f)


def load_members():
    global face_members
    if os.path.exists(FACE_DB_FILE):
        try:
            with open(FACE_DB_FILE, "rb") as f:
                face_members = pickle.load(f)
        except Exception as exc:
            face_members = {}
            add_log(f"⚠️ 人脸数据库读取失败: {exc}")
    return face_members


def save_members():
    with open(FACE_DB_FILE, "wb") as f:
        pickle.dump(face_members, f)


def current_time_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_time(value):
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return None


def re_search(pattern, text):
    return re.search(pattern, str(text or "")) is not None


def mark_state_dirty():
    global state_dirty
    state_dirty = True


def sync_runtime_globals(service_instance):
    global TOLERANCE, ALARM_INTERVAL, LOG_LIMIT, CAMERA_INDEX, PUSH_ENABLED
    settings = service_instance.settings
    TOLERANCE = float(settings.get("tolerance", TOLERANCE))
    ALARM_INTERVAL = int(settings.get("alarm_interval", ALARM_INTERVAL))
    LOG_LIMIT = int(settings.get("log_limit", LOG_LIMIT))
    CAMERA_INDEX = int(settings.get("camera_index", CAMERA_INDEX))
    PUSH_ENABLED = bool(settings.get("push_enabled", PUSH_ENABLED))
    BARK_KEYS[1] = settings.get("bark_key", BARK_KEYS[1])
    PUSH_CONFIG[1]["key"] = BARK_KEYS[1]


def clean_record_list(records, limit):
    if not isinstance(records, list):
        return []
    return [item for item in records if isinstance(item, dict)][:limit]


def build_persistent_state(service_instance):
    sensor_keys = [key for key, *_ in SENSOR_DEFINITIONS]
    return {
        "version": 1,
        "saved_at": current_time_text(),
        "settings": dict(getattr(service_instance, "settings", {})),
        "thresholds": dict(getattr(service_instance, "thresholds", DEFAULT_THRESHOLDS)),
        "system_logs": list(system_logs[:LOG_LIMIT]),
        "sensor_history": list(sensor_history[:SENSOR_HISTORY_LIMIT]),
        "sensor_values": {key: monitor_data.get(key) for key in sensor_keys},
        "monitor_data": {
            "alert_history": list(monitor_data["alert_history"]),
            "action_history": list(monitor_data["action_history"]),
            "stranger_alarm_count": int(monitor_data.get("stranger_alarm_count", 0)),
            "last_stranger_time": monitor_data.get("last_stranger_time", "--"),
            "last_face_event": monitor_data.get("last_face_event", "暂无人脸事件")
        }
    }


def save_app_state(force=False):
    global state_dirty, last_state_save
    now = time.time()
    with data_lock:
        if not force and not state_dirty:
            return
        if not force and now - last_state_save < STATE_SAVE_INTERVAL:
            return
        payload = build_persistent_state(globals().get("service"))
        state_dirty = False
        last_state_save = now
    tmp_file = f"{APP_STATE_FILE}.tmp"
    try:
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp_file, APP_STATE_FILE)
    except Exception as exc:
        with data_lock:
            state_dirty = True
        safe_print(f"状态文件保存失败: {exc}")


def load_app_state(service_instance):
    global system_logs, sensor_history, state_dirty, last_state_save
    if not os.path.exists(APP_STATE_FILE):
        return
    try:
        with open(APP_STATE_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception as exc:
        safe_print(f"状态文件读取失败: {exc}")
        return

    restored_safe_values = False
    with data_lock:
        settings = payload.get("settings")
        if isinstance(settings, dict):
            service_instance.settings.update({
                key: value for key, value in settings.items()
                if key in service_instance.settings
            })
        thresholds = payload.get("thresholds")
        if isinstance(thresholds, dict):
            for key in DEFAULT_THRESHOLDS:
                if key in thresholds:
                    try:
                        service_instance.thresholds[key] = float(thresholds[key])
                    except (TypeError, ValueError):
                        pass
        sync_runtime_globals(service_instance)
        service_instance.sync_threshold_rules()

        system_logs[:] = clean_record_list(payload.get("system_logs"), LOG_LIMIT)
        sensor_history[:] = clean_record_list(payload.get("sensor_history"), SENSOR_HISTORY_LIMIT)

        persisted_monitor = payload.get("monitor_data", {})
        if isinstance(persisted_monitor, dict):
            monitor_data["alert_history"] = clean_record_list(persisted_monitor.get("alert_history"), 300)[:60]
            monitor_data["action_history"] = clean_record_list(persisted_monitor.get("action_history"), 160)[:40]
            monitor_data["stranger_alarm_count"] = int(persisted_monitor.get("stranger_alarm_count", 0) or 0)
            monitor_data["last_stranger_time"] = persisted_monitor.get("last_stranger_time", "--") or "--"
            monitor_data["last_face_event"] = persisted_monitor.get("last_face_event", "暂无人脸事件") or "暂无人脸事件"

        sensor_values = payload.get("sensor_values")
        if isinstance(sensor_values, dict):
            sensor_keys = [key for key, *_ in SENSOR_DEFINITIONS]
            candidate = {key: monitor_data[key] for key in sensor_keys}
            for key in sensor_keys:
                if key in sensor_values:
                    try:
                        candidate[key] = float(sensor_values[key])
                    except (TypeError, ValueError):
                        pass
            if evaluate_accident(candidate) is None:
                monitor_data.update(candidate)
                restored_safe_values = True

        service_instance.stranger_alarm_count = int(monitor_data.get("stranger_alarm_count", 0))
        service_instance.last_stranger_time = monitor_data.get("last_stranger_time", "--")
        service_instance.last_face_event = monitor_data.get("last_face_event", "暂无人脸事件")
        state_dirty = False
        last_state_save = time.time()

    restored_text = "已恢复上次安全范围内的监测快照" if restored_safe_values else "已恢复历史记录与系统设置"
    safe_print(restored_text)


def extract_face_feature(face_gray):
    if cv2 is None or np is None or face_gray is None or face_gray.size == 0:
        return None
    resized = cv2.resize(cv2.equalizeHist(face_gray), FACE_FEATURE_SIZE)
    normalized = resized.astype("float32") / 255.0
    normalized = (normalized - float(normalized.mean())) / (float(normalized.std()) + 1e-6)
    thumbnail = cv2.resize(normalized, (32, 32)).flatten().astype("float32")

    center = resized[1:-1, 1:-1]
    lbp = np.zeros_like(center, dtype=np.uint8)
    lbp |= (resized[:-2, :-2] >= center).astype(np.uint8) << 7
    lbp |= (resized[:-2, 1:-1] >= center).astype(np.uint8) << 6
    lbp |= (resized[:-2, 2:] >= center).astype(np.uint8) << 5
    lbp |= (resized[1:-1, 2:] >= center).astype(np.uint8) << 4
    lbp |= (resized[2:, 2:] >= center).astype(np.uint8) << 3
    lbp |= (resized[2:, 1:-1] >= center).astype(np.uint8) << 2
    lbp |= (resized[2:, :-2] >= center).astype(np.uint8) << 1
    lbp |= (resized[1:-1, :-2] >= center).astype(np.uint8)
    texture_hist = cv2.calcHist([lbp], [0], None, [256], [0, 256]).flatten().astype("float32")
    texture_hist /= float(texture_hist.sum()) + 1e-6

    grad_x = cv2.Sobel(resized, cv2.CV_32F, 1, 0, ksize=3)
    grad_y = cv2.Sobel(resized, cv2.CV_32F, 0, 1, ksize=3)
    magnitude, angle = cv2.cartToPolar(grad_x, grad_y, angleInDegrees=True)
    bins = (angle / 22.5).astype("int32") % 16
    edge_hist = np.zeros(16, dtype="float32")
    for bucket in range(16):
        edge_hist[bucket] = float(magnitude[bins == bucket].sum())
    edge_hist /= float(edge_hist.sum()) + 1e-6

    feature = np.concatenate([
        thumbnail * 0.55,
        texture_hist * 0.35,
        edge_hist * 0.10
    ]).astype("float32")
    return normalize_face_feature(feature)


def normalize_face_feature(feature):
    if np is None or feature is None:
        return None
    arr = np.asarray(feature, dtype="float32").flatten()
    if arr.size != FACE_FEATURE_LENGTH:
        return None
    norm = float(np.linalg.norm(arr))
    if norm <= 1e-6:
        return None
    return arr / norm


def compare_faces(feature1, feature2):
    current = normalize_face_feature(feature1)
    saved = normalize_face_feature(feature2)
    if current is None or saved is None:
        return None
    return max(0.0, min(2.0, 1.0 - float(np.dot(current, saved))))


def face_match_threshold():
    try:
        value = float(TOLERANCE)
    except (TypeError, ValueError):
        return FACE_MATCH_THRESHOLD
    if 0 < value < 1:
        return min(max(value, 0.20), 0.50)
    if 1 <= value <= 100:
        return min(max(0.20 + value * 0.002, 0.22), 0.42)
    return FACE_MATCH_THRESHOLD


def identify_face(face_feature):
    if cv2 is None or face_feature is None or not face_members:
        return None, 0
    best_name = None
    best_score = 999.0
    second_score = 999.0
    threshold = face_match_threshold()
    for name, features in face_members.items():
        person_score = 999.0
        for saved in features:
            score = compare_faces(face_feature, saved)
            if score is None:
                continue
            if score < person_score:
                person_score = score
        if person_score < best_score:
            second_score = best_score
            best_score = person_score
            best_name = name
        elif person_score < second_score:
            second_score = person_score
    is_clear_match = second_score >= 999.0 or (second_score - best_score) >= FACE_AMBIGUITY_MARGIN
    if best_name and best_score <= threshold and is_clear_match:
        return best_name, best_score
    return None, best_score


def match_conditions(sensors, conditions):
    for param, threshold in conditions.items():
        value = sensors.get(param, 0)
        if "min" in threshold and value < threshold["min"]:
            return False
        if "max" in threshold and value > threshold["max"]:
            return False
    return True


def evaluate_accident(sensors):
    matched = []
    for key, rule in ACCIDENT_RULES.items():
        if match_conditions(sensors, rule["conditions"]):
            matched.append((rule["level"], key, rule))
    if not matched:
        return None
    matched.sort(key=lambda item: item[0], reverse=True)
    return matched[0]


class LabSafetyService:
    def __init__(self):
        self.running = True
        self.start_time = datetime.now()
        self.current_level = 0
        self.current_accident = "无"
        self.last_sensor_snapshot = []
        self.last_history_time = 0
        self.last_push_time = 0
        self.camera = None
        self.camera_running = False
        self.camera_source = "none"
        self.camera_lock = threading.RLock()
        self.camera_thread = None
        self.face_cascade = None
        self.latest_frame = None
        self.latest_face_feature = None
        self.current_face_count = 0
        self.known_face_count = 0
        self.unknown_face_count = 0
        self.detected_faces = []
        self.last_face_event = "暂无人脸事件"
        self.last_stranger_time = "--"
        self.stranger_alarm_count = 0
        self.camera_error = ""
        self.settings = {
            "camera_index": CAMERA_INDEX,
            "tolerance": TOLERANCE,
            "alarm_interval": ALARM_INTERVAL,
            "log_limit": LOG_LIMIT,
            "push_enabled": PUSH_ENABLED,
            "bark_key": BARK_KEYS[1],
            "auto_start_camera": False,
            "save_intruder_snapshot": True,
            "face_sample_limit": 3,
            "device_overrides": {},
            "system_name": "智巡卫士",
            "lab_location": "实验室",
            "refresh_interval": 1
        }
        self.thresholds = DEFAULT_THRESHOLDS.copy()

    def start(self):
        global service_started
        service_started = True
        load_app_state(self)
        load_user_database()
        load_members()
        add_log("🚀 Flask Web 监控服务已启动")
        threading.Thread(target=self.update_data_loop, daemon=True).start()
        threading.Thread(target=self.analyze_risk_loop, daemon=True).start()

    def update_data_loop(self):
        while self.running:
            with data_lock:
                if self.current_level == 0:
                    monitor_data["room_temp"] = max(22, min(monitor_data["room_temp"] + random.uniform(-0.25, 0.25), 29))
                    monitor_data["humidity"] = max(35, min(monitor_data["humidity"] + random.uniform(-1.5, 1.5), 65))
                    monitor_data["smoke"] = max(0, min(monitor_data["smoke"] + random.randint(-1, 1), 8))
                    monitor_data["voc"] = max(30, min(monitor_data["voc"] + random.randint(-4, 4), 120))
                    monitor_data["gas_ch4"] = max(0, min(monitor_data["gas_ch4"] + random.randint(-30, 30), 400))
                    monitor_data["gas_h2"] = max(0, min(monitor_data["gas_h2"] + random.randint(-20, 20), 300))
                    monitor_data["gas_co"] = max(0, min(monitor_data["gas_co"] + random.randint(-1, 1), 10))
                    monitor_data["gas_h2s"] = max(0, min(monitor_data["gas_h2s"] + random.randint(-1, 1), 3))
                    monitor_data["reactor_temp"] = max(35, min(monitor_data["reactor_temp"] + random.uniform(-0.35, 0.35), 55))
                    monitor_data["reactor_pressure"] = max(0.3, min(monitor_data["reactor_pressure"] + random.uniform(-0.015, 0.015), 0.7))
                monitor_data["face_count"] = len(face_members)
                monitor_data["current_faces"] = self.current_face_count
                monitor_data["known_face_count"] = self.known_face_count
                monitor_data["unknown_face_count"] = self.unknown_face_count
                monitor_data["detected_faces"] = list(self.detected_faces)
                monitor_data["last_face_event"] = self.last_face_event
                monitor_data["last_stranger_time"] = self.last_stranger_time
                monitor_data["stranger_alarm_count"] = self.stranger_alarm_count
                monitor_data["camera_running"] = self.camera_running
                monitor_data["camera_source"] = self.camera_source
                monitor_data["camera_error"] = self.camera_error
            self.record_sensor_history()
            save_app_state()
            time.sleep(1)

    def record_sensor_history(self):
        now = time.time()
        if now - self.last_history_time < SENSOR_HISTORY_INTERVAL:
            return
        rows = self.sensor_rows()
        with data_lock:
            sensor_history.insert(0, {
                "time": current_time_text(),
                "risk_level": monitor_data.get("risk_level", 0),
                "accident_type": monitor_data.get("accident_type", "无"),
                "rows": rows
            })
            del sensor_history[SENSOR_HISTORY_LIMIT:]
            mark_state_dirty()
        self.last_history_time = now

    def analyze_risk_loop(self):
        while self.running:
            with data_lock:
                sensors = {key: monitor_data[key] for key in (
                    "room_temp", "smoke", "voc", "gas_ch4", "gas_h2", "gas_co", "gas_h2s",
                    "reactor_temp", "reactor_pressure"
                )}
            accident_info = evaluate_accident(sensors)
            if accident_info:
                level, _key, rule = accident_info
                self.set_risk(level, rule["name"], rule["actions"], rule.get("fire_suppression"), rule.get("sequence", []))
            else:
                self.set_risk(0, "无", [], None, NORMAL_SEQUENCE)
            time.sleep(1)

    def set_risk(self, level, accident_name, actions, fire_suppression, sequence):
        if level == self.current_level and accident_name == self.current_accident:
            return
        self.current_level = level
        self.current_accident = accident_name
        with data_lock:
            if level == 0:
                monitor_data.update({
                    "risk_level": 0,
                    "risk_text": "✅ 系统正常",
                    "risk_reason": "所有参数正常",
                    "accident_type": "无",
                    "alert_detail": "",
                    "actions": [],
                    "emergency_sequence": NORMAL_SEQUENCE[:],
                    "fire_suppression": None
                })
                mark_state_dirty()
                return

            level_text = {1: "📢 一级预警", 2: "⚠️ 二级报警", 3: "🚨 三级紧急"}[level]
            action_display = actions[:]
            if fire_suppression:
                action_display = [fire_suppression] + action_display
            monitor_data.update({
                "risk_level": level,
                "risk_text": level_text,
                "risk_reason": f"{accident_name} 触发",
                "accident_type": accident_name,
                "alert_detail": f"{accident_name}触发，执行联动",
                "actions": action_display,
                "emergency_sequence": sequence,
                "fire_suppression": fire_suppression
            })
            history_entry = {
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "level": level,
                "level_text": level_text,
                "type": accident_name,
                "message": f"{accident_name}触发，执行联动"
            }
            monitor_data["alert_history"].insert(0, history_entry)
            del monitor_data["alert_history"][60:]
            for action in actions:
                monitor_data["action_history"].insert(0, {
                    "time": datetime.now().strftime("%H:%M:%S"),
                    "action": action
                })
            del monitor_data["action_history"][40:]
            mark_state_dirty()
        add_log(f"【{level_text}】{accident_name}")
        self.send_bark_push(level, level_text, accident_name)

    def send_bark_push(self, level, title, message, log_missing=False):
        if not self.settings["push_enabled"]:
            if log_missing:
                add_log("⚠️ Bark 推送未启用，陌生人报警已记录在系统")
            return False
        config = PUSH_CONFIG.get(level, PUSH_CONFIG[1])
        url = self.build_bark_url(title, message)
        if not url:
            if log_missing:
                add_log("⚠️ Bark Key 未配置，陌生人报警已记录在系统")
            return False

        def push_worker():
            try:
                response = requests.get(url, timeout=5)
                response.raise_for_status()
                add_log(f"📱 推送({level}级): {', '.join(config['targets'])}")
            except Exception as exc:
                add_log(f"⚠️ 推送失败: {str(exc)[:40]}")

        threading.Thread(target=push_worker, daemon=True).start()
        return True

    def build_bark_url(self, title, message, bark_key=None):
        key_or_url = (bark_key if bark_key is not None else self.settings.get("bark_key", "")).strip()
        if not key_or_url:
            return None
        base = key_or_url.rstrip("/")
        if not base.startswith(("http://", "https://")):
            base = f"https://api.day.app/{base}"
        encoded_title = urllib.parse.quote(str(title))
        encoded_message = urllib.parse.quote(
            f"{message}\n时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        query = urllib.parse.urlencode({
            "sound": "alarm",
            "group": self.settings.get("system_name", "实验室安全")
        })
        return f"{base}/{encoded_title}/{encoded_message}?{query}"

    def test_bark_push(self, bark_key=None, enabled=True):
        if not enabled:
            return False, "请先勾选启用 Bark 手机推送"
        url = self.build_bark_url("智巡卫士测试推送", "Bark 推送配置成功", bark_key)
        if not url:
            return False, "请填写 Bark Key 或 Bark URL"
        try:
            response = requests.get(url, timeout=8)
            response.raise_for_status()
            add_log("📱 Bark 测试推送已发送")
            return True, "测试推送已发送，请查看 iPhone Bark"
        except Exception as exc:
            return False, f"Bark 推送失败: {str(exc)[:80]}"

    def start_camera(self):
        if cv2 is None:
            self.camera_error = "未安装 OpenCV，无法打开摄像头"
            add_log("⚠️ 未安装 OpenCV，无法打开摄像头")
            return False, self.camera_error
        if self.camera_running:
            return True, "摄像头已运行"

        index = int(self.settings["camera_index"])
        cap = cv2.VideoCapture(index, cv2.CAP_DSHOW)
        if not cap.isOpened():
            cap = cv2.VideoCapture(index, cv2.CAP_ANY)
        if not cap.isOpened():
            self.camera_error = "摄像头打开失败，请检查设备权限或编号"
            add_log(f"⚠️ 摄像头打开失败，编号: {index}")
            return False, self.camera_error

        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        self.camera = cap
        self.camera_running = True
        self.camera_source = "server"
        self.camera_error = ""
        self.camera_thread = threading.Thread(target=self.camera_loop, daemon=True)
        self.camera_thread.start()
        add_log("📷 摄像头已打开")
        return True, "摄像头已打开"

    def start_client_camera(self):
        if cv2 is None or np is None:
            self.camera_error = "服务器未安装 OpenCV，无法识别人脸"
            add_log("⚠️ 服务器未安装 OpenCV，无法识别人脸")
            return False, self.camera_error
        with self.camera_lock:
            if self.camera is not None:
                self.camera.release()
                self.camera = None
            self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
            self.camera_running = True
            self.camera_source = "browser"
            self.camera_error = ""
            self.latest_frame = None
            self.latest_face_feature = None
            self.current_face_count = 0
            self.known_face_count = 0
            self.unknown_face_count = 0
            self.detected_faces = []
            self.last_face_event = "浏览器摄像头已连接，等待画面"
        add_log("📷 浏览器端摄像头已连接")
        return True, "浏览器摄像头已连接"

    def stop_camera(self):
        self.camera_running = False
        with self.camera_lock:
            if self.camera is not None:
                self.camera.release()
                self.camera = None
            self.camera_source = "none"
            self.latest_frame = None
            self.latest_face_feature = None
            self.current_face_count = 0
            self.known_face_count = 0
            self.unknown_face_count = 0
            self.detected_faces = []
            self.last_face_event = "摄像头已关闭"
        add_log("📷 摄像头已关闭")
        return True, "摄像头已关闭"

    def process_client_frame(self, frame_bytes, include_image=True):
        if cv2 is None or np is None:
            return False, "服务器未安装 OpenCV", None, 0, 0
        if not self.camera_running or self.camera_source != "browser":
            ok, message = self.start_client_camera()
            if not ok:
                return False, message, None, 0, 0
        image_array = np.frombuffer(frame_bytes, dtype=np.uint8)
        frame = cv2.imdecode(image_array, cv2.IMREAD_COLOR)
        if frame is None:
            return False, "浏览器画面解码失败", None, 0, 0
        processed = self.process_face_frame(frame)
        with self.camera_lock:
            self.latest_frame = processed
        height, width = processed.shape[:2]
        if not include_image:
            return True, "ok", None, width, height
        ok, buffer = cv2.imencode(".jpg", processed, [int(cv2.IMWRITE_JPEG_QUALITY), 70])
        if not ok:
            return False, "识别画面编码失败", None, width, height
        image_data = base64.b64encode(buffer.tobytes()).decode("ascii")
        return True, "ok", f"data:image/jpeg;base64,{image_data}", width, height

    def camera_loop(self):
        while self.camera_running:
            with self.camera_lock:
                cap = self.camera
            if cap is None:
                time.sleep(0.05)
                continue
            ret, frame = cap.read()
            if not ret:
                self.camera_error = "读取摄像头画面失败"
                time.sleep(0.1)
                continue
            processed = self.process_face_frame(frame)
            with self.camera_lock:
                self.latest_frame = processed
            time.sleep(0.03)

    @staticmethod
    def expand_face_box(x, y, w, h, frame_width, frame_height):
        pad_x = int(w * 0.24)
        pad_top = int(h * 0.30)
        pad_bottom = int(h * 0.22)
        left = max(0, x - pad_x)
        top = max(0, y - pad_top)
        right = min(frame_width, x + w + pad_x)
        bottom = min(frame_height, y + h + pad_bottom)
        return left, top, max(1, right - left), max(1, bottom - top)

    def process_face_frame(self, frame):
        if cv2 is None:
            return frame
        frame_height, frame_width = frame.shape[:2]
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.1, 5) if self.face_cascade is not None else []
        self.current_face_count = len(faces)
        self.known_face_count = 0
        self.unknown_face_count = 0
        detected_faces = []
        self.latest_face_feature = None

        for index, (x, y, w, h) in enumerate(faces):
            face_roi = gray[y:y + h, x:x + w]
            face_feature = extract_face_feature(face_roi)
            box_x, box_y, box_w, box_h = self.expand_face_box(x, y, w, h, frame_width, frame_height)
            if index == 0 and face_feature is not None:
                self.latest_face_feature = face_feature.copy()
            name, _score = identify_face(face_feature)
            if name:
                color = (0, 220, 70)
                label = name
                status = "已授权"
                status_code = "known"
                self.known_face_count += 1
            else:
                color = (0, 0, 255)
                label = "STRANGER"
                status = "陌生人"
                status_code = "unknown"
                self.unknown_face_count += 1
                self.handle_stranger_alarm(x, y, frame, (box_x, box_y, box_w, box_h))
            detected_faces.append({
                "label": label,
                "status": status,
                "status_code": status_code,
                "score": None if _score is None or _score >= 999 else round(float(_score), 4),
                "box": [int(box_x), int(box_y), int(box_w), int(box_h)]
            })
            cv2.rectangle(frame, (box_x, box_y), (box_x + box_w, box_y + box_h), color, 2)
            cv2.putText(frame, label, (box_x, max(box_y - 8, 18)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

        info = f"Members: {len(face_members)}  Faces: {self.current_face_count}"
        cv2.putText(frame, info, (10, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 2)
        self.detected_faces = detected_faces
        if self.current_face_count:
            self.last_face_event = f"检测到 {self.current_face_count} 张人脸，陌生人 {self.unknown_face_count} 张"
        else:
            self.last_face_event = "画面内暂无人脸"
        return frame

    def save_stranger_snapshot(self, frame, box):
        if cv2 is None or frame is None or not self.settings.get("save_intruder_snapshot", True):
            return None
        try:
            os.makedirs(SNAPSHOT_DIR, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
            filename = f"stranger_{timestamp}.jpg"
            path = os.path.join(SNAPSHOT_DIR, filename)
            preview = frame.copy()
            x, y, w, h = [int(value) for value in box]
            cv2.rectangle(preview, (x, y), (x + w, y + h), (0, 0, 255), 2)
            cv2.imwrite(path, preview, [int(cv2.IMWRITE_JPEG_QUALITY), 82])

            snapshots = sorted(
                (os.path.join(SNAPSHOT_DIR, item) for item in os.listdir(SNAPSHOT_DIR) if item.lower().endswith(".jpg")),
                key=os.path.getmtime,
                reverse=True
            )
            for old_file in snapshots[200:]:
                try:
                    os.remove(old_file)
                except OSError:
                    pass
            return f"/static/snapshots/{filename}"
        except Exception as exc:
            add_log(f"⚠️ 陌生人抓拍保存失败: {str(exc)[:40]}")
            return None

    def handle_stranger_alarm(self, x, y, frame=None, box=None):
        alarm_key = f"{x // 50}_{y // 50}"
        now = time.time()
        interval = int(self.settings["alarm_interval"])
        if alarm_key in last_alarm_time and (now - last_alarm_time[alarm_key]) <= interval:
            return
        last_alarm_time[alarm_key] = now
        snapshot_url = self.save_stranger_snapshot(frame, box or (x, y, 1, 1))
        entry = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "level": 1,
            "level_text": "📢 一级预警",
            "type": "陌生人闯入",
            "message": "人脸识别检测到陌生人"
        }
        if snapshot_url:
            entry["snapshot_url"] = snapshot_url
        with data_lock:
            monitor_data["alert_history"].insert(0, entry)
            del monitor_data["alert_history"][60:]
            monitor_data["stranger_alarm_count"] += 1
            mark_state_dirty()
        self.stranger_alarm_count += 1
        self.last_stranger_time = entry["time"]
        self.last_face_event = "检测到陌生人并生成报警"
        add_log("🚨 人脸识别: 检测到陌生人")
        self.send_bark_push(1, "👤 陌生人闯入", "实验室检测到陌生人", log_missing=True)

    def frame_generator(self):
        if cv2 is None:
            return
        while True:
            with self.camera_lock:
                frame = None if self.latest_frame is None else self.latest_frame.copy()
            if frame is None:
                frame = self.placeholder_frame()
            ok, buffer = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 72])
            if ok:
                yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + buffer.tobytes() + b"\r\n"
            time.sleep(0.06)

    def placeholder_frame(self):
        frame = cv2.imread(os.path.join(BASE_DIR, "static", "img", "camera-placeholder.jpg")) if cv2 else None
        if frame is not None:
            return frame
        frame = 255 * (0 * __import__("numpy").ones((360, 640, 3), dtype="uint8"))
        cv2.putText(frame, "Camera Offline", (190, 175), cv2.FONT_HERSHEY_SIMPLEX, 1, (160, 170, 190), 2)
        return frame

    def add_current_face_member(self, name):
        if not self.camera_running:
            return False, "请先打开摄像头"
        if not name:
            return False, "姓名不能为空"
        if self.latest_face_feature is None:
            return False, "当前画面未检测到可录入的人脸"
        with data_lock:
            current_samples = [
                feature for feature in face_members.get(name, [])
                if normalize_face_feature(feature) is not None
            ]
            current_samples.append(self.latest_face_feature.copy())
            sample_limit = int(self.settings.get("face_sample_limit", 3))
            face_members[name] = current_samples[-sample_limit:]
            save_members()
        add_log(f"👤 人脸录入: {name}")
        return True, f"已录入 {name}"

    def delete_face_member(self, name):
        if name not in face_members:
            return False, "人员不存在"
        with data_lock:
            face_members.pop(name, None)
            save_members()
        add_log(f"🗑️ 删除人脸库人员: {name}")
        return True, "删除成功"

    def apply_settings(self, payload):
        global TOLERANCE, ALARM_INTERVAL, LOG_LIMIT
        try:
            camera_index = int(payload.get("camera_index", self.settings["camera_index"]))
            tolerance = float(payload.get("tolerance", self.settings["tolerance"]))
            alarm_interval = int(payload.get("alarm_interval", self.settings["alarm_interval"]))
            log_limit = int(payload.get("log_limit", self.settings["log_limit"]))
            face_sample_limit = int(payload.get("face_sample_limit", self.settings["face_sample_limit"]))
            refresh_interval = int(payload.get("refresh_interval", self.settings["refresh_interval"]))
        except ValueError:
            return False, "设置参数必须是数字"
        if camera_index < 0 or not 1 <= tolerance <= 100 or alarm_interval < 1 or log_limit < 20 or face_sample_limit < 1 or refresh_interval < 1:
            return False, "请检查参数范围"

        self.settings.update({
            "camera_index": camera_index,
            "tolerance": tolerance,
            "alarm_interval": alarm_interval,
            "log_limit": log_limit,
            "push_enabled": bool(payload.get("push_enabled", False)),
            "auto_start_camera": bool(payload.get("auto_start_camera", False)),
            "save_intruder_snapshot": bool(payload.get("save_intruder_snapshot", False)),
            "face_sample_limit": face_sample_limit,
            "system_name": payload.get("system_name", self.settings["system_name"]).strip() or "智巡卫士",
            "lab_location": payload.get("lab_location", self.settings["lab_location"]).strip() or "实验室",
            "refresh_interval": refresh_interval,
            "bark_key": payload.get("bark_key", "")
        })
        for key in DEFAULT_THRESHOLDS:
            if key in payload:
                try:
                    self.thresholds[key] = float(payload[key])
                except (TypeError, ValueError):
                    return False, f"{key} 阈值必须是数字"
        TOLERANCE = tolerance
        ALARM_INTERVAL = alarm_interval
        LOG_LIMIT = log_limit
        self.sync_threshold_rules()
        BARK_KEYS[1] = self.settings["bark_key"]
        PUSH_CONFIG[1]["key"] = BARK_KEYS[1]
        with data_lock:
            del system_logs[LOG_LIMIT:]
        add_log("⚙️ 系统设置已保存")
        return True, "设置已保存"

    def sync_threshold_rules(self):
        ACCIDENT_RULES["high_temp_warning"]["conditions"]["room_temp"]["min"] = self.thresholds["room_temp_warn"]
        ACCIDENT_RULES["indoor_fire"]["conditions"]["smoke"]["min"] = self.thresholds["smoke_alarm"]
        ACCIDENT_RULES["gas_ch4_leak"]["conditions"]["gas_ch4"]["min"] = self.thresholds["gas_ch4_warn"]
        ACCIDENT_RULES["gas_ch4_explosion"]["conditions"]["gas_ch4"]["min"] = self.thresholds["gas_ch4_explosion"]
        ACCIDENT_RULES["co_leak"]["conditions"]["gas_co"]["min"] = self.thresholds["gas_co_alarm"]
        ACCIDENT_RULES["h2s_leak"]["conditions"]["gas_h2s"]["min"] = self.thresholds["gas_h2s_alarm"]
        ACCIDENT_RULES["reactor_fire"]["conditions"]["reactor_temp"]["min"] = self.thresholds["reactor_temp_alarm"]
        ACCIDENT_RULES["reactor_overpressure"]["conditions"]["reactor_pressure"]["min"] = self.thresholds["reactor_pressure_alarm"]

    def sensor_rows(self):
        rows = []
        with data_lock:
            for key, name, unit, max_value, threshold_key in SENSOR_DEFINITIONS:
                value = monitor_data[key]
                threshold = self.thresholds.get(threshold_key) if threshold_key else None
                ratio = 0 if max_value == 0 else min(100, max(0, (float(value) / max_value) * 100))
                status = "正常"
                if threshold is not None and value >= threshold:
                    status = "报警" if "alarm" in threshold_key or "explosion" in threshold_key else "预警"
                rows.append({
                    "key": key,
                    "name": name,
                    "value": round(value, 2),
                    "unit": unit,
                    "threshold": threshold,
                    "status": status,
                    "ratio": round(ratio, 1)
                })
        self.last_sensor_snapshot = rows
        return rows

    def device_rows(self):
        with data_lock:
            actions = " ".join(monitor_data.get("actions", []))
            level = int(monitor_data.get("risk_level", 0) or 0)
            fire_suppression = monitor_data.get("fire_suppression")
            camera_running = bool(monitor_data.get("camera_running"))
            camera_source = monitor_data.get("camera_source", "none")
        push_ready = bool(self.settings.get("push_enabled") and self.settings.get("bark_key"))
        overrides = self.settings.get("device_overrides", {})
        if not isinstance(overrides, dict):
            overrides = {}

        dynamic = {
            "sensor_hub": ("在线", "online", "每 30 秒记录历史快照"),
            "ventilation": (
                "已联动" if re_search("排风|通风", actions) else "待命",
                "active" if re_search("排风|通风", actions) else "online",
                "风险等级触发时自动排风"
            ),
            "suppression": (
                "已启用" if fire_suppression or re_search("灭火|抑制|冷却", actions) else "待命",
                "danger" if level >= 3 and (fire_suppression or re_search("灭火|抑制|冷却", actions)) else "online",
                fire_suppression or "火灾/过压时自动联动"
            ),
            "access_control": (
                "疏散联动" if re_search("门禁|安全门|疏散", actions) else "正常",
                "warning" if re_search("门禁|安全门|疏散", actions) else "online",
                "紧急时自动打开安全通道"
            ),
            "camera": (
                "运行中" if camera_running else "未启动",
                "active" if camera_running else "standby",
                f"{'浏览器摄像头' if camera_source == 'browser' else '服务器摄像头'}" if camera_running else "等待开启"
            ),
            "bark_push": (
                "已配置" if push_ready else "未配置",
                "active" if push_ready else "warning",
                "陌生人和安全报警会推送 iPhone" if push_ready else "请在系统设置填写 Bark Key"
            )
        }

        devices = []
        for item in DEVICE_CATALOG:
            override = overrides.get(item["id"], {}) if isinstance(overrides.get(item["id"], {}), dict) else {}
            enabled = bool(override.get("enabled", True))
            maintenance = bool(override.get("maintenance", False))
            status, status_code, note = dynamic[item["id"]]
            if not enabled:
                status, status_code, note = "停用", "offline", "设备已手动停用"
            elif maintenance:
                status, status_code, note = "维护中", "warning", "设备处于维护标记"
            devices.append({
                **item,
                "enabled": enabled,
                "maintenance": maintenance,
                "status": status,
                "status_code": status_code,
                "note": note,
                "updated_at": override.get("updated_at", "--")
            })
        return devices

    def update_device(self, payload):
        device_id = payload.get("id")
        action = payload.get("action")
        if device_id not in {item["id"] for item in DEVICE_CATALOG}:
            return False, "设备不存在"
        overrides = self.settings.setdefault("device_overrides", {})
        if not isinstance(overrides, dict):
            overrides = {}
            self.settings["device_overrides"] = overrides
        current = overrides.setdefault(device_id, {})
        if action == "toggle":
            current["enabled"] = not bool(current.get("enabled", True))
        elif action == "maintenance":
            current["maintenance"] = not bool(current.get("maintenance", False))
        elif action == "enable":
            current["enabled"] = True
            current["maintenance"] = False
        else:
            return False, "未知设备操作"
        current["updated_at"] = current_time_text()
        mark_state_dirty()
        device_name = next(item["name"] for item in DEVICE_CATALOG if item["id"] == device_id)
        add_log(f"🛠️ 设备管理: {device_name} {action}")
        return True, "设备状态已更新"

    def report_payload(self, period):
        days = 7 if period == "week" else 30
        label = "周报" if period == "week" else "月报"
        cutoff = datetime.now() - timedelta(days=days)
        with data_lock:
            history = list(sensor_history)
            alerts = list(monitor_data["alert_history"])
        snapshots = []
        for item in history:
            item_time = parse_time(item.get("time"))
            if item_time is None or item_time >= cutoff:
                snapshots.append(item)
        if not snapshots:
            snapshots = [{
                "time": current_time_text(),
                "risk_level": monitor_data.get("risk_level", 0),
                "accident_type": monitor_data.get("accident_type", "无"),
                "rows": self.sensor_rows()
            }]

        summary = []
        detail = []
        for key, name, unit, _max_value, _threshold_key in SENSOR_DEFINITIONS:
            values = []
            exceed_count = 0
            latest = None
            for snapshot in snapshots:
                for row in snapshot.get("rows", []):
                    if row.get("key") != key:
                        continue
                    value = float(row.get("value", 0))
                    values.append(value)
                    if latest is None:
                        latest = row
                    if row.get("status") != "正常":
                        exceed_count += 1
                    detail.append({
                        "时间": snapshot.get("time", "--"),
                        "监测项": row.get("name", name),
                        "数值": row.get("value", ""),
                        "单位": row.get("unit", unit),
                        "状态": row.get("status", "正常"),
                        "风险等级": snapshot.get("risk_level", 0),
                        "事故类型": snapshot.get("accident_type", "无")
                    })
            if values:
                summary.append({
                    "监测项": name,
                    "单位": unit,
                    "最新值": latest.get("value", "") if latest else "",
                    "平均值": round(sum(values) / len(values), 2),
                    "最高值": round(max(values), 2),
                    "最低值": round(min(values), 2),
                    "异常次数": exceed_count,
                    "样本数": len(values)
                })
        alert_count = sum(1 for item in alerts if (parse_time(item.get("time")) or datetime.now()) >= cutoff)
        return {
            "period": period,
            "label": label,
            "days": days,
            "generated_at": current_time_text(),
            "summary": summary,
            "detail": detail,
            "alert_count": alert_count,
            "snapshot_count": len(snapshots)
        }

    def export_rows(self):
        snapshot = self.snapshot()
        return [
            {
                "监测项": row["name"],
                "当前值": row["value"],
                "单位": row["unit"],
                "阈值": "" if row["threshold"] is None else row["threshold"],
                "状态": row["status"],
                "归一化百分比": row["ratio"],
                "风险等级": snapshot["risk_level"],
                "事故类型": snapshot["accident_type"],
                "导出时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            for row in snapshot["sensor_rows"]
        ]

    def snapshot(self):
        with data_lock:
            data = dict(monitor_data)
            data["alert_history"] = list(monitor_data["alert_history"])
            data["action_history"] = list(monitor_data["action_history"])
            data["sensor_history_count"] = len(sensor_history)
            data["latest_sensor_history"] = list(sensor_history[:20])
            data["trend_history"] = list(sensor_history[:120])
        runtime = datetime.now() - self.start_time
        data["runtime"] = {
            "hours": int(runtime.total_seconds() // 3600),
            "minutes": int((runtime.total_seconds() % 3600) // 60)
        }
        data["members"] = len(face_members)
        data["users"] = len(user_database)
        data["logs"] = len(system_logs)
        data["thresholds"] = dict(self.thresholds)
        data["settings"] = dict(self.settings)
        data["sensor_rows"] = self.sensor_rows()
        data["devices"] = self.device_rows()
        data["server_time"] = current_time_text()
        data["last_state_saved"] = datetime.fromtimestamp(last_state_save).strftime("%Y-%m-%d %H:%M:%S") if last_state_save else "--"
        return data


service = LabSafetyService()


def flush_app_state_on_exit():
    if not service_started:
        return
    try:
        save_app_state(force=True)
    except Exception:
        pass


atexit.register(flush_app_state_on_exit)


def login_required(handler):
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return jsonify({"success": False, "message": "Unauthorized"}), 401
        return handler(*args, **kwargs)
    wrapper.__name__ = handler.__name__
    return wrapper


@web_app.route("/")
def index():
    if "username" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html", asset_version=ASSET_VERSION)


@web_app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("index"))
    return render_template(
        "dashboard.html",
        username=session["username"],
        role=session.get("role", "实验人员"),
        asset_version=ASSET_VERSION
    )


@web_app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    password_hash = hashlib.md5(password.encode()).hexdigest()
    if username in user_database and user_database[username]["password"] == password_hash:
        session["username"] = username
        session["role"] = user_database[username].get("role", "实验人员")
        add_log(f"🔐 用户登录: {username}")
        if service.settings["auto_start_camera"]:
            add_log("ℹ️ Web 版自动开摄像头需由浏览器授权，登录后请点击打开摄像头")
        return jsonify({"success": True, "redirect": url_for("dashboard")})
    return jsonify({"success": False, "message": "用户名或密码错误"})


@web_app.route("/api/register", methods=["POST"])
def api_register():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    email = data.get("email", "").strip()
    if len(username) < 3:
        return jsonify({"success": False, "message": "用户名至少 3 位"})
    if len(password) < 6:
        return jsonify({"success": False, "message": "密码至少 6 位"})
    if username in user_database:
        return jsonify({"success": False, "message": "用户名已存在"})
    user_database[username] = {
        "password": hashlib.md5(password.encode()).hexdigest(),
        "email": email or "-",
        "role": "实验人员",
        "display_name": username,
        "department": "-",
        "phone": "-",
        "register_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    save_user_database()
    add_log(f"📝 新用户注册: {username}")
    return jsonify({"success": True, "message": "注册成功"})


@web_app.route("/api/logout", methods=["POST"])
@login_required
def api_logout():
    username = session.get("username")
    session.clear()
    add_log(f"🔒 用户退出: {username}")
    return jsonify({"success": True, "redirect": url_for("index")})


@web_app.route("/api/data")
@login_required
def api_data():
    return jsonify(service.snapshot())


@web_app.route("/api/logs", methods=["GET", "DELETE"])
@login_required
def api_logs():
    if request.method == "DELETE":
        with data_lock:
            system_logs.clear()
        add_log("🧹 系统日志已清空")
    return jsonify({"success": True, "logs": system_logs})


@web_app.route("/api/users", methods=["GET", "POST", "PATCH", "DELETE"])
@login_required
def api_users():
    if request.method in ("POST", "PATCH"):
        data = request.get_json() or {}
        username = data.get("username", "").strip()
        if len(username) < 3:
            return jsonify({"success": False, "message": "用户名至少 3 位"})
        is_new = request.method == "POST" or username not in user_database
        password = data.get("password", "")
        if is_new and len(password) < 6:
            return jsonify({"success": False, "message": "新增用户密码至少 6 位"})
        if username == "admin" and data.get("role") != "实验室管理员":
            return jsonify({"success": False, "message": "默认管理员角色不能降级"})
        user_database.setdefault(username, {
            "password": hashlib.md5(password.encode()).hexdigest(),
            "register_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        info = user_database[username]
        if password:
            if len(password) < 6:
                return jsonify({"success": False, "message": "密码至少 6 位"})
            info["password"] = hashlib.md5(password.encode()).hexdigest()
        info["role"] = data.get("role") or info.get("role", "实验人员")
        info["email"] = data.get("email", "-").strip() or "-"
        info["display_name"] = data.get("display_name", username).strip() or username
        info["department"] = data.get("department", "-").strip() or "-"
        info["phone"] = data.get("phone", "-").strip() or "-"
        info.setdefault("register_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        save_user_database()
        add_log(f"{'➕ 新增' if is_new else '✏️ 更新'}用户: {username}")

    if request.method == "DELETE":
        data = request.get_json() or {}
        username = data.get("username", "")
        if username == "admin":
            return jsonify({"success": False, "message": "默认管理员不能删除"})
        if username == session.get("username"):
            return jsonify({"success": False, "message": "不能删除当前登录用户"})
        if username not in user_database:
            return jsonify({"success": False, "message": "用户不存在"})
        user_database.pop(username, None)
        save_user_database()
        add_log(f"🗑️ 删除用户: {username}")
    users = [
        {
            "username": username,
            "role": info.get("role", "-"),
            "email": info.get("email", "-"),
            "display_name": info.get("display_name", username),
            "department": info.get("department", "-"),
            "phone": info.get("phone", "-"),
            "register_time": info.get("register_time", "-")
        }
        for username, info in sorted(user_database.items())
    ]
    return jsonify({"success": True, "users": users})


@web_app.route("/api/face-members", methods=["GET", "POST", "DELETE"])
@login_required
def api_face_members():
    if request.method == "POST":
        data = request.get_json() or {}
        success, message = service.add_current_face_member(data.get("name", "").strip())
        return jsonify({"success": success, "message": message})
    if request.method == "DELETE":
        data = request.get_json() or {}
        success, message = service.delete_face_member(data.get("name", ""))
        return jsonify({"success": success, "message": message})
    members = []
    for name, features in sorted(face_members.items()):
        valid_samples = sum(1 for feature in features if normalize_face_feature(feature) is not None)
        legacy_samples = max(0, len(features) - valid_samples)
        if valid_samples >= 3:
            status = "样本充足"
        elif valid_samples > 0:
            status = "可继续补录"
        else:
            status = "旧样本需重新录入"
        if legacy_samples:
            status = f"{status} · {legacy_samples} 个旧样本已忽略"
        members.append({
            "name": name,
            "samples": valid_samples,
            "total_samples": len(features),
            "legacy_samples": legacy_samples,
            "status": status
        })
    return jsonify({"success": True, "members": members})


@web_app.route("/api/camera/start", methods=["POST"])
@login_required
def api_camera_start():
    success, message = service.start_camera()
    return jsonify({"success": success, "message": message})


@web_app.route("/api/client-camera/start", methods=["POST"])
@login_required
def api_client_camera_start():
    success, message = service.start_client_camera()
    return jsonify({"success": success, "message": message})


@web_app.route("/api/camera/stop", methods=["POST"])
@login_required
def api_camera_stop():
    success, message = service.stop_camera()
    return jsonify({"success": success, "message": message})


@web_app.route("/api/client-frame", methods=["POST"])
@login_required
def api_client_frame():
    frame_file = request.files.get("frame")
    if frame_file is None:
        return jsonify({"success": False, "message": "未收到摄像头画面"}), 400
    include_image = request.args.get("preview") != "0"
    success, message, image, frame_width, frame_height = service.process_client_frame(frame_file.read(), include_image)
    payload = {
        "success": success,
        "message": message,
        "frame_width": frame_width,
        "frame_height": frame_height
    }
    if image:
        payload["image"] = image
    if success:
        payload["faces"] = service.detected_faces
        payload["current_faces"] = service.current_face_count
        payload["known_face_count"] = service.known_face_count
        payload["unknown_face_count"] = service.unknown_face_count
    return jsonify(payload)


@web_app.route("/video_feed")
def video_feed():
    if "username" not in session:
        return Response(status=401)
    if cv2 is None:
        return Response(status=503)
    return Response(service.frame_generator(), mimetype="multipart/x-mixed-replace; boundary=frame")


@web_app.route("/api/settings", methods=["GET", "POST"])
@login_required
def api_settings():
    if request.method == "POST":
        success, message = service.apply_settings(request.get_json() or {})
        return jsonify({"success": success, "message": message, "settings": service.settings, "thresholds": service.thresholds})
    return jsonify({"success": True, "settings": service.settings, "thresholds": service.thresholds})


@web_app.route("/api/push/test", methods=["POST"])
@login_required
def api_push_test():
    payload = request.get_json() or {}
    success, message = service.test_bark_push(
        payload.get("bark_key", service.settings.get("bark_key", "")),
        bool(payload.get("push_enabled", service.settings.get("push_enabled", False)))
    )
    return jsonify({"success": success, "message": message})


@web_app.route("/api/devices", methods=["GET", "POST"])
@login_required
def api_devices():
    if request.method == "POST":
        success, message = service.update_device(request.get_json() or {})
        return jsonify({"success": success, "message": message, "devices": service.device_rows()})
    return jsonify({"success": True, "devices": service.device_rows()})


@web_app.route("/api/report/<period>/<fmt>")
@login_required
def api_report(period, fmt):
    if period not in {"week", "month"}:
        return jsonify({"success": False, "message": "不支持的报表周期"}), 400
    report = service.report_payload(period)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename_prefix = f"lab_safety_{period}_report_{timestamp}"
    add_log(f"📄 导出{report['label']}: {fmt.upper()}")

    if fmt == "json":
        payload = json.dumps(report, ensure_ascii=False, indent=2)
        return Response(
            payload,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.json"}
        )

    if fmt == "csv":
        buffer = io.StringIO()
        buffer.write(f"{report['label']},生成时间,{report['generated_at']},历史快照,{report['snapshot_count']},报警数,{report['alert_count']}\n")
        rows = report["summary"]
        writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        return Response(
            "\ufeff" + buffer.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"}
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return jsonify({"success": False, "message": "当前环境缺少 openpyxl，无法导出 Excel"}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = report["label"]
        meta_rows = [
            ["报表周期", report["label"]],
            ["生成时间", report["generated_at"]],
            ["历史快照", report["snapshot_count"]],
            ["报警数量", report["alert_count"]]
        ]
        for row in meta_rows:
            ws.append(row)
        ws.append([])
        headers = list(report["summary"][0].keys())
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for row in report["summary"]:
            ws.append([row[h] for h in headers])

        detail_ws = wb.create_sheet("历史明细")
        detail_headers = list(report["detail"][0].keys())
        detail_ws.append(detail_headers)
        for cell in detail_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for row in report["detail"]:
            detail_ws.append([row[h] for h in detail_headers])

        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells) + 2
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 30)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{filename_prefix}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return jsonify({"success": False, "message": "不支持的报表格式"}), 400


@web_app.route("/api/export/<fmt>")
@login_required
def api_export(fmt):
    rows = service.export_rows()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    add_log(f"📤 导出监测状态: {fmt.upper()}")

    if fmt == "json":
        payload = json.dumps(rows, ensure_ascii=False, indent=2)
        return Response(
            payload,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename=monitor_status_{timestamp}.json"}
        )

    if fmt == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        return Response(
            "\ufeff" + buffer.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=monitor_status_{timestamp}.csv"}
        )

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except Exception:
            return jsonify({"success": False, "message": "当前环境缺少 openpyxl，无法导出 Excel"}), 500

        wb = Workbook()
        ws = wb.active
        ws.title = "监测状态"
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for row in rows:
            ws.append([row[h] for h in headers])
        for column_cells in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 28)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=f"monitor_status_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return jsonify({"success": False, "message": "不支持的导出格式"}), 400


@web_app.route("/camera_offline.svg")
def camera_offline():
    svg = """<svg xmlns="http://www.w3.org/2000/svg" width="960" height="600" viewBox="0 0 960 600">
<rect width="960" height="600" fill="#05080c"/>
<rect x="28" y="28" width="904" height="544" fill="none" stroke="#1f2b38" stroke-width="2"/>
<text x="480" y="286" text-anchor="middle" fill="#607080" font-family="Microsoft YaHei,Arial" font-size="34">摄像头已关闭</text>
<text x="480" y="330" text-anchor="middle" fill="#3f5062" font-family="Arial" font-size="18">Camera Offline</text>
</svg>"""
    return Response(svg, mimetype="image/svg+xml")


@web_app.route("/api/simulate/<mode>", methods=["POST"])
@login_required
def api_simulate(mode):
    message = "操作已执行"
    accident_info = None
    with data_lock:
        if mode == "fire":
            monitor_data["smoke"] = 35
            add_log("🔥 模拟室内火灾: 烟雾35%")
            message = "已触发室内火灾模拟"
        elif mode == "reactor":
            monitor_data["reactor_temp"] = 95
            add_log("⚗️ 模拟反应釜火灾: 釜温95℃")
            message = "已触发反应釜火灾模拟"
        elif mode == "gas":
            monitor_data["gas_ch4"] = 8000
            add_log("💨 模拟瓦斯泄漏: CH₄浓度8000ppm")
            message = "已触发瓦斯泄漏模拟"
        elif mode == "reset":
            monitor_data.update({
                "room_temp": 25.0, "humidity": 45, "smoke": 0, "voc": 50,
                "reactor_temp": 45, "reactor_pressure": 0.5,
                "gas_ch4": 0, "gas_h2": 0, "gas_co": 0, "gas_h2s": 0
            })
            add_log("✅ 系统已恢复正常")
            message = "系统已恢复正常"
        else:
            return jsonify({"success": False, "message": "未知模拟类型"}), 400

        sensors = {key: monitor_data[key] for key in (
            "room_temp", "smoke", "voc", "gas_ch4", "gas_h2", "gas_co", "gas_h2s",
            "reactor_temp", "reactor_pressure"
        )}
        accident_info = evaluate_accident(sensors)
        service.current_level = -1
        service.current_accident = "__simulation__"

    if accident_info:
        level, _key, rule = accident_info
        service.set_risk(level, rule["name"], rule["actions"], rule.get("fire_suppression"), rule.get("sequence", []))
    else:
        service.set_risk(0, "无", [], None, NORMAL_SEQUENCE)

    return jsonify({
        "success": True,
        "message": message,
        "risk_level": monitor_data.get("risk_level", 0),
        "accident_type": monitor_data.get("accident_type", "无")
    })


def main():
    service.start()
    safe_print("")
    safe_print("=" * 60)
    safe_print("Zhixun Weishi Web server started")
    safe_print("Open: http://127.0.0.1:5000")
    safe_print("Default account: admin / admin123")
    safe_print("=" * 60)
    safe_print("")
    web_app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        with open(os.path.join(BASE_DIR, "startup_error.log"), "w", encoding="utf-8") as f:
            f.write(traceback.format_exc())
        raise
